import os
import re
from flask import Flask, render_template, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from dotenv import load_dotenv
import pandas as pd
from datetime import datetime
from typing import List, Dict
from urllib.parse import quote

from scraper.sirene import SireneClient

app = Flask(__name__)

# Configuration de la base de données
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL') or \
    'sqlite:///' + os.path.join(basedir, 'newbiz.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Modèle pour stocker les données des entreprises
class EntrepriseData(db.Model):
    __tablename__ = 'entreprise_data'
    
    siret = db.Column(db.String(14), primary_key=True)
    statut = db.Column(db.String(50), default='A traiter')
    date_modification = db.Column(db.String(50))
    funbooster = db.Column(db.Text, default='')
    observation = db.Column(db.Text, default='')
    
    def to_dict(self):
        return {
            'siret': self.siret,
            'statut': self.statut or 'A traiter',
            'date_modification': self.date_modification or '',
            'funbooster': self.funbooster or '',
            'observation': self.observation or ''
        }

# Créer les tables au démarrage
with app.app_context():
    db.create_all()

# Charger la clé API
load_dotenv()
api_key = os.getenv("SIRENE_API_KEY")
client = SireneClient(api_key=api_key)


def generate_pappers_url(siren: str) -> str:
    """Génère une URL de recherche Pappers pour trouver le dirigeant."""
    if not siren or len(siren) < 9:
        return ""
    return f"https://www.pappers.fr/recherche?q={siren}"


def generate_pagesjaunes_url(nom: str, adresse: str) -> str:
    """
    Génère une URL PagesJaunes pour trouver le téléphone de l'entreprise.
    Format : https://www.pagesjaunes.fr/recherche/{code_postal}/{nom_entreprise}
    - code_postal : extrait de l'adresse (5 chiffres)
    - nom_entreprise : nom encodé pour l'URL
    """
    if not nom:
        return ""

    code_postal = ""
    if adresse:
        match = re.search(r"\b(\d{5})\b", adresse)
        if match:
            code_postal = match.group(1)

    if not code_postal:
        # Sans code postal, le lien serait moins précis ; on préfère ne rien mettre
        return ""

    encoded_nom = quote(nom.strip())
    return f"https://www.pagesjaunes.fr/recherche/{code_postal}/{encoded_nom}"


def generate_opco_url(siret: str) -> str:
    """
    Génère un lien vers le site 'Quel est mon OPCO ?' de France Compétences.
    On passe le SIRET en paramètre d'URL (s'il est valide) pour faciliter la saisie.
    """
    if not siret:
        return ""
    siret = str(siret).strip()
    if not (siret.isdigit() and len(siret) == 14):
        return ""
    # Même si le site n'exploite pas encore ce paramètre, ça permet au téléconseiller
    # de voir le SIRET dans l'URL et de le copier/coller facilement.
    return f"https://quel-est-mon-opco.francecompetences.fr/?siret={siret}"


@app.route('/')
def index():
    """Page d'accueil avec l'interface de recherche."""
    return render_template('index.html')


@app.route('/api/search', methods=['POST'])
def search_companies():
    """API endpoint pour rechercher des entreprises."""
    try:
        data = request.json
        secteur = data.get('secteur', '').strip()
        departement = data.get('departement', '').strip()
        
        if not secteur or not departement:
            return jsonify({'error': 'Veuillez remplir les champs Secteur et Département.'}), 400
        
        # Lancer la recherche
        results = client.search_by_secteur_and_departement(
            secteur=secteur,
            departement=departement,
            limit=300,
        )
        
        # Filtrer uniquement les entreprises avec l'état "Actif"
        active_results = [ent for ent in results if str(ent.get("etat", "")).strip() == "Actif"]
        
        # Ajouter les liens Pappers (dirigeant), PagesJaunes (téléphone)
        # et France Compétences (OPCO) à chaque résultat
        # Et récupérer les données sauvegardées depuis la base de données
        for ent in active_results:
            siren = ent.get("siren", "")
            nom = ent.get("nom", "")
            adresse = ent.get("adresse", "")
            siret = ent.get("siret", "")
            
            ent["pappers_url"] = generate_pappers_url(siren)
            ent["pagesjaunes_url"] = generate_pagesjaunes_url(nom, adresse)
            ent["opco_url"] = generate_opco_url(siret)
            
            # Récupérer les données depuis la base de données
            if siret:
                entreprise_data = EntrepriseData.query.filter_by(siret=siret).first()
                if entreprise_data:
                    ent["statut"] = entreprise_data.statut or 'A traiter'
                    ent["date_modification"] = entreprise_data.date_modification or ''
                    ent["funbooster"] = entreprise_data.funbooster or ''
                    ent["observation"] = entreprise_data.observation or ''
                else:
                    ent["statut"] = 'A traiter'
                    ent["date_modification"] = ''
                    ent["funbooster"] = ''
                    ent["observation"] = ''
        
        return jsonify({
            'success': True,
            'count': len(active_results),
            'results': active_results
        })
    
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la recherche : {str(e)}'}), 500


@app.route('/api/save-statut', methods=['POST'])
def save_statut():
    """API endpoint pour sauvegarder le statut d'une entreprise."""
    try:
        data = request.json
        siret = data.get('siret', '').strip()
        statut = data.get('statut', 'A traiter').strip()
        
        if not siret:
            return jsonify({'error': 'SIRET manquant.'}), 400
        
        # Créer ou mettre à jour l'enregistrement
        entreprise_data = EntrepriseData.query.filter_by(siret=siret).first()
        
        if not entreprise_data:
            entreprise_data = EntrepriseData(siret=siret)
            db.session.add(entreprise_data)
        
        entreprise_data.statut = statut
        entreprise_data.date_modification = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'date_modification': entreprise_data.date_modification
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Erreur lors de la sauvegarde : {str(e)}'}), 500


@app.route('/api/save-field', methods=['POST'])
def save_field():
    """API endpoint pour sauvegarder FunBooster ou Observation."""
    try:
        data = request.json
        siret = data.get('siret', '').strip()
        field = data.get('field', '').strip()  # 'funbooster' ou 'observation'
        value = data.get('value', '').strip()
        
        if not siret or field not in ['funbooster', 'observation']:
            return jsonify({'error': 'Paramètres invalides.'}), 400
        
        # Créer ou mettre à jour l'enregistrement
        entreprise_data = EntrepriseData.query.filter_by(siret=siret).first()
        
        if not entreprise_data:
            entreprise_data = EntrepriseData(siret=siret)
            db.session.add(entreprise_data)
        
        setattr(entreprise_data, field, value)
        
        db.session.commit()
        
        return jsonify({'success': True})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Erreur lors de la sauvegarde : {str(e)}'}), 500


@app.route('/api/export', methods=['POST'])
def export_to_excel():
    """API endpoint pour exporter les résultats en Excel."""
    try:
        data = request.json
        results = data.get('results', [])
        
        if not results:
            return jsonify({'error': 'Aucune donnée à exporter.'}), 400
        
        # Définir l'ordre exact des colonnes
        column_order = [
            "Nom",
            "Adresse",
            "Téléphone",
            "Secteur",
            "SIRET",
            "SIREN",
            "Effectif",
            "État",
            "Statut",
            "Date de modification",
            "FunBooster",
            "Observation",
            "Lien OPCO (France Compétences)",
            "Lien Dirigeant (Pappers)",
            "Lien Téléphone (PagesJaunes)",
        ]
        
        # Filtrer uniquement les entreprises avec l'état "Actif"
        active_results = [ent for ent in results if str(ent.get("etat", "")).strip() == "Actif"]
        
        if not active_results:
            return jsonify({'error': 'Aucune entreprise active à exporter.'}), 400
        
        # Récupérer les données depuis la base de données pour chaque entreprise
        for ent in active_results:
            siret = ent.get("siret", "").strip()
            if siret:
                entreprise_data = EntrepriseData.query.filter_by(siret=siret).first()
                if entreprise_data:
                    ent["statut"] = entreprise_data.statut or 'A traiter'
                    ent["date_modification"] = entreprise_data.date_modification or ''
                    ent["funbooster"] = entreprise_data.funbooster or ''
                    ent["observation"] = entreprise_data.observation or ''
        
        # Préparer les données pour Excel avec nettoyage
        excel_data = []
        for ent in active_results:
            # Nettoyer et préparer chaque valeur
            nom = str(ent.get("nom", "")).strip() if ent.get("nom") else ""
            adresse = str(ent.get("adresse", "")).strip() if ent.get("adresse") else ""
            telephone = str(ent.get("telephone", "")).strip() if ent.get("telephone") else ""
            secteur = str(ent.get("secteur", "")).strip() if ent.get("secteur") else ""
            siret = str(ent.get("siret", "")).strip() if ent.get("siret") else ""
            siren = str(ent.get("siren", "")).strip() if ent.get("siren") else ""
            effectif = str(ent.get("effectif", "")).strip() if ent.get("effectif") else ""
            etat = str(ent.get("etat", "")).strip() if ent.get("etat") else ""
            statut = str(ent.get("statut", "A traiter")).strip() if ent.get("statut") else "A traiter"
            date_modification = str(ent.get("date_modification", "")).strip() if ent.get("date_modification") else ""
            funbooster = str(ent.get("funbooster", "")).strip() if ent.get("funbooster") else ""
            observation = str(ent.get("observation", "")).strip() if ent.get("observation") else ""
            opco_url = str(ent.get("opco_url", "")).strip() if ent.get("opco_url") else ""
            pappers_url = str(ent.get("pappers_url", "")).strip() if ent.get("pappers_url") else ""
            pagesjaunes_url = str(ent.get("pagesjaunes_url", "")).strip() if ent.get("pagesjaunes_url") else ""
            
            excel_data.append({
                "Nom": nom,
                "Adresse": adresse,
                "Téléphone": telephone,
                "Secteur": secteur,
                "SIRET": siret,
                "SIREN": siren,
                "Effectif": effectif,
                "État": etat,
                "Statut": statut,
                "Date de modification": date_modification,
                "FunBooster": funbooster,
                "Observation": observation,
                "Lien OPCO (France Compétences)": opco_url,
                "Lien Dirigeant (Pappers)": pappers_url,
                "Lien Téléphone (PagesJaunes)": pagesjaunes_url,
            })
        
        # Créer un DataFrame avec l'ordre des colonnes spécifié
        df = pd.DataFrame(excel_data, columns=column_order)
        filename = f"entreprises_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('temp', filename)
        
        # Créer le dossier temp s'il n'existe pas
        os.makedirs('temp', exist_ok=True)
        
        # Exporter vers Excel avec formatage
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Créer le fichier Excel
        df.to_excel(filepath, index=False, engine='openpyxl')
        
        # Améliorer le formatage du fichier Excel
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Style pour les en-têtes
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Appliquer le style aux en-têtes
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Ajuster la largeur des colonnes automatiquement
        column_widths = {
            'A': 30,  # Nom
            'B': 40,  # Adresse
            'C': 20,  # Téléphone
            'D': 20,  # Secteur
            'E': 18,  # SIRET
            'F': 15,  # SIREN
            'G': 15,  # Effectif
            'H': 15,  # État
            'I': 20,  # Statut
            'J': 25,  # Date de modification
            'K': 20,  # FunBooster
            'L': 30,  # Observation
            'M': 40,  # Lien OPCO (France Compétences)
            'N': 50,  # Lien Dirigeant (Pappers)
            'O': 50,  # Lien Téléphone (PagesJaunes)
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Alignement du contenu (aligné à gauche, centré verticalement)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Enregistrer le fichier
        wb.save(filepath)
        
        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'error': f'Erreur lors de l\'export : {str(e)}'}), 500


# Créer les dossiers nécessaires au démarrage
os.makedirs('temp', exist_ok=True)

if __name__ == '__main__':
    print("\n" + "="*50)
    print("🚀 Serveur web démarré !")
    print("📱 Ouvrez votre navigateur et allez sur :")
    print("   http://127.0.0.1:5000")
    print("="*50 + "\n")
    
    # En production, Render utilise gunicorn, donc on ne lance pas app.run()
    # En développement local, on peut lancer avec app.run()
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)

